# --------------------------     Birthday Wisher      ---------------
# -------------------------- Developed by Github: @satya-root ---------------

import openpyxl
from twilio.rest import Client
import time
import datetime
import sys
import os

# For Windows -> \\ -------- Example -> C:\\Users\\Downloads\\SMS-BirthdayWisher\\DB.xlsx
# For Linux   -> /  -------- Example -> /home/kali/Downloads/SMS-BirthdayWisher/DB.xlsx
workbook = openpyxl.load_workbook('------Location of DB.xlsx------') 
active = workbook.active

def convert_to_secs():
    current_hr = int(time.strftime("%H"))       
    current_min = int(time.strftime("%M"))
    current_secs = int(time.strftime("%S"))

    # -----   Total 86400 secs in 1 day     -----
    # -----       1 hr = 3600secs           -----
    # -----   12:00 a.m. = 00:00 hours      -----
    secs = (current_hr * 3600) + (current_min * 60) + (current_secs)
    remaining_secs = 86400 - secs
    print(current_hr)
    print(current_min)
    return remaining_secs


date = datetime.date.today()
str_date = str(date.day) + '_0' + str(date.month)

z = convert_to_secs()
print('Remaining Secs till 12 A.M. = ', z)

if (z != 0):
    # ----- Twilio Declarations -----
    account_sid = '--- ACCOUNT SID ---'
    auth_token = '--- AUTH TOKEN ---'
    client = Client(account_sid, auth_token)
    time.sleep(z)

    # ----- Excel Sheet Iteration -----
    for i in range(2, active.max_row+1):
        birthday = str(active.cell(row=i, column=2).value)

        if (str_date == birthday):
            ph = str(int(active.cell(row=i, column=3).value))
            wishes = str(active.cell(row=i, column=4).value)
            # ----- Uncomment next lines to see your list of phone numbers and wishes -----
            # print('+91' + ph)
            # print(wishes)

            # ----- Send Message -----
            message = client.messages.create(
                from_='+12058929610',
                body=wishes,
                to='+91' + ph
            )
            print(message.sid)

        # If you want the program to reexecute itself when it stops, then uncomment both of the following lines.
        # rerun = sys.executable    
        # os.execl(rerun, rerun, * sys.argv)
        



# Credits: https://github.com/satya-root/
# -------------------- End -----------------------
